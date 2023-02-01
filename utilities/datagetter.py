from openpyxl import load_workbook
from utilities.confreader import confread

def getdata(excelpath, sheetname):
    workbook=load_workbook(confread('basic_info',excelpath))
    sheet=workbook[confread('basic_info',sheetname)]
    total_rows=sheet.max_row
    total_cols=sheet.max_column

    main=[]
    for i in range(2,total_rows+1):
        temp=[]
        for j in range(1,total_cols+1):
            temp.insert(j,sheet.cell(row=i, column=j).value)
        main.insert(i,temp)
    return main